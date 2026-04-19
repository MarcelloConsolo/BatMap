import json
import os
import shutil
from openpyxl import load_workbook

def fix_excel_files():
    years = [2022, 2023, 2024, 2025]
    json_path = 'app/src/main/assets/comuni_italiani.json'

    if not os.path.exists(json_path):
        print(f"Errore: {json_path} non trovato.")
        return

    # Carica database comuni
    with open(json_path, encoding='utf-8') as f:
        db_comuni = json.load(f)

    for year in years:
        file_name = f'Pipistrelli {year}.xlsx'
        if not os.path.exists(file_name):
            print(f"Salto {file_name}: non trovato.")
            continue

        print(f"--- Elaborazione {file_name} ---")
        wb = load_workbook(file_name)
        ws = wb.active

        # Trova intestazioni
        header_row = 1
        col_map = {}
        for r in range(1, 10):
            row_vals = [str(ws.cell(row=r, column=c).value).lower() for c in range(1, ws.max_column + 1)]
            if any('specie' in x or 'comune' in x for x in row_vals):
                header_row = r
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=r, column=c).value).lower()
                    if 'specie' in val: col_map['specie'] = c
                    if 'comune' in val or 'citt' in val: col_map['comune'] = c
                    if 'prov' in val or val == 'pr': col_map['prov'] = c
                    if 'localit' in val or 'indirizzo' in val: col_map['loc'] = c
                    if 'lat' in val: col_map['lat'] = c
                    if 'lon' in val or 'lng' in val: col_map['lon'] = c
                break

        if 'comune' not in col_map:
            print(f"Errore: colonna Comune non trovata in {file_name}")
            continue

        count_prov = 0
        count_coords = 0
        for r in range(header_row + 1, ws.max_row + 1):
            comune = str(ws.cell(row=r, column=col_map['comune']).value or "").strip().lower()

            if comune in db_comuni:
                info = db_comuni[comune]

                # FORZA la provincia corretta se quella attuale è mancante o errata
                if 'prov' in col_map:
                    curr_prov = str(ws.cell(row=r, column=col_map['prov']).value or "").strip().upper()
                    if curr_prov != info['prov'].upper():
                        ws.cell(row=r, column=col_map['prov']).value = info['prov'].upper()
                        count_prov += 1

                # VERIFICA COORDINATE: Se mancano o sono palesemente 0, le ripristiniamo dal DB del comune
                if 'lat' in col_map:
                    lat_val = ws.cell(row=r, column=col_map['lat']).value
                    if lat_val is None or lat_val == 0 or str(lat_val).strip() == "":
                        ws.cell(row=r, column=col_map['lat']).value = info['lat']
                        count_coords += 1
                if 'lon' in col_map:
                    lon_val = ws.cell(row=r, column=col_map['lon']).value
                    if lon_val is None or lon_val == 0 or str(lon_val).strip() == "":
                        ws.cell(row=r, column=col_map['lon']).value = info['lon']
                        count_coords += 1

        wb.save(file_name)
        # Copia anche in assets e web
        shutil.copy(file_name, f'app/src/main/assets/{file_name}')
        if os.path.exists('web'):
            shutil.copy(file_name, f'web/{file_name}')

        print(f"Completato {file_name}: aggiornate {count_prov} province e {count_coords} coordinate.")

if __name__ == "__main__":
    fix_excel_files()
